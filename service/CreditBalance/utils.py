from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import pandas as pd
from pprint import pprint


def generate_credit_balance_report(df: pd.DataFrame):
    org_col_names = df.columns.tolist()
    converted_col_names = [
        "海纜名稱",
        "海纜作業",
        "記帳段號",
        "Invoice No.",
        "Issue Date",
        "CN no.",
        "Issue Date",
        "Description",
        "Debit",
        "Credit",
        "Balance",
    ]
    col_names_map = dict(zip(org_col_names, converted_col_names))

    # start to generate excel from openpyxl
    max_col = len(converted_col_names)
    max_row = len(df.index) + 2  # include header and comment(Currency:USD, col names)

    wb = Workbook()
    ws = wb.active
    ws.title = "Credit Balance Report"

    # set column width
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 20
    ws.column_dimensions["K"].width = 20

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            # ws.cell(row=row, column=col).border = Border(
            #     left=Side(border_style="thin", color="FF000000"),
            #     right=Side(border_style="thin", color="FF000000"),
            #     top=Side(border_style="thin", color="FF000000"),
            #     bottom=Side(border_style="thin", color="FF000000"),
            # )
            if row == 1:
                # no borderline(transparent)
                ws.cell(row=row, column=col).border = Border(
                    left=Side(border_style="thin", color="00FFFFFF"),
                    right=Side(border_style="thin", color="00FFFFFF"),
                    top=Side(border_style="thin", color="00FFFFFF"),
                    bottom=Side(border_style="thin", color="00FFFFFF"),
                )
                if col == 11:
                    # set comment
                    ws.cell(row=row, column=col).value = "Currency: USD"
            elif row == 2:
                # set header
                ws.cell(row=row, column=col).value = converted_col_names[col - 1]
                ws.cell(row=row, column=col).font = Font(bold=True)
                ws.cell(row=row, column=col).alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color="FFC000", end_color="FFC000", fill_type="solid"
                )
            elif row > 2:
                # set data
                ws.cell(row=row, column=col).value = df.iloc[row - 3, col - 1]
                ws.cell(row=row, column=col).alignment = Alignment(
                    horizontal="center", vertical="center"
                )

    file_name = f"CB歷程.xlsx"
    wb.save(file_name)


if __name__ == "__main__":
    df = pd.read_excel("CreditBalanceReport_20230417142422.xlsx")
    generate_credit_balance_report(df)
