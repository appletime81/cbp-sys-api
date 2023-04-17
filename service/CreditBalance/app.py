import os

from fastapi import APIRouter, Request, status, Depends
from crud import *
from get_db import get_db
from sqlalchemy.orm import Session
from utils.utils import *
from utils.orm_pydantic_convert import *
from copy import deepcopy
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from fastapi.responses import StreamingResponse, FileResponse

router = APIRouter()


@router.get("/CreditBalance/{urlCondition}", status_code=status.HTTP_200_OK)
async def getCreditBalance(
    request: Request,
    urlCondition: str,
    db: Session = Depends(get_db),
):
    crud = CRUD(db, CreditBalanceDBModel)
    table_name = "CB"
    if "getByBillDetail=yes" not in urlCondition:
        if urlCondition == "all":
            CreditBalanceDataList = crud.get_all()
        elif "start" in urlCondition or "end" in urlCondition:
            dictCondition = convert_url_condition_to_dict(urlCondition)
            sql_condition = convert_dict_to_sql_condition(dictCondition, table_name)

            # get all CreditBalance by sql
            CreditBalanceDataList = crud.get_all_by_sql(sql_condition)
        else:
            dictCondition = convert_url_condition_to_dict(urlCondition)
            CreditBalanceDataList = crud.get_with_condition(dictCondition)
    else:
        # urlCondition: SubmarineCable=str&WorkTitle=str&BillMilestone=str&PartyName=str
        urlCondition = urlCondition.replace("getByBillDetail=yes", "")
        dictCondition = convert_url_condition_to_dict(urlCondition)
        CreditBalanceDataList = crud.get_with_condition(dictCondition)
    return CreditBalanceDataList


@router.post("/CreditBalanceStatement", status_code=status.HTTP_200_OK)
async def getCreditBalanceStatement(
    request: Request,
    urlCondition: str,
    db: Session = Depends(get_db),
):
    """
    {"CBID": int}
    """
    crud = CRUD(db, CreditBalanceStatementDBModel)
    CBID = (await request.json())["CBID"]
    CBStatementDataList = crud.get_with_condition({"CBID": CBID})
    return CBStatementDataList


# @router.post("/CreditBalanceStatement/generateReport")
# async def getCreditBalanceStatement(
#     request: Request,
#     CBIDPydanticData: CBIDSchema,
#     db: Session = Depends(get_db),
# ):
#     """
#     {"CBID": int}
#     """
#
#     crudCreditBalanceStatement = CRUD(db, CreditBalanceStatementDBModel)
#     crudCreditBalance = CRUD(db, CreditBalanceDBModel)
#
#     CBID = (await request.json())["CBID"]
#     CBStatementDataList = crudCreditBalanceStatement.get_with_condition({"CBID": CBID})
#     CBData = crudCreditBalance.get_with_condition({"CBID": CBID})[0]
#
#     # use openpyxl to generate report
#
#     workbook = Workbook()
#
#     # generate col title
#     worksheet = workbook.active
#     # 隱藏所有格線
#     worksheet.sheet_view.showGridLines = False
#     worksheet.title = "CB歷程"
#     col_names = ["BM", "Ref No.", "Description", "Debit", "Credit", "Balance"]
#     OriginalCreditBalance = CBData.CurrAmount
#
#     for tempCBStatementData in CBStatementDataList[::-1]:
#         OriginalCreditBalance -= tempCBStatementData.TransAmount
#
#     # generate dict 1-7 mapping to alphabet A-G
#     alphabetDict = {}
#     for i in range(1, 8):
#         alphabetDict[i] = chr(i + 64)
#     max_rows = len(CBStatementDataList) + 1
#     max_cols = 6
#     for i in range(1, max_rows + 2):
#         if i == 2:
#             value_dict = {
#                 1: CBData.BillMilestone,
#                 2: CBData.BillingNo,
#                 3: CBData.Note,
#                 4: "",
#                 5: OriginalCreditBalance,
#                 6: OriginalCreditBalance,
#             }
#         elif i > 2:
#             value_dict = {1: CBData.BillMilestone, 2: CBData.BillingNo, 3: CBData.Note}
#             OriginalCreditBalance += CBStatementDataList[i - 3].TransAmount
#             value_dict[6] = OriginalCreditBalance
#             if CBStatementDataList[i - 3].TransAmount > 0:
#                 value_dict[4] = ""
#                 value_dict[5] = CBStatementDataList[i - 3].TransAmount
#             else:
#                 value_dict[4] = abs(CBStatementDataList[i - 3].TransAmount)
#                 value_dict[5] = ""
#         for j in range(1, max_cols + 1):
#             worksheet[f"{alphabetDict[j]}{i}"].alignment = Alignment(
#                 horizontal="center", vertical="center"
#             )
#             worksheet[f"{alphabetDict[j]}{i}"].font = Font(name="微軟正黑體", size=12)
#             # 設為粗體字
#             worksheet[f"{alphabetDict[j]}{i}"].font = Font(bold=True)
#             worksheet[f"{alphabetDict[j]}{i}"].border = Border(
#                 left=Side(border_style="thin", color="FF000000"),
#                 right=Side(border_style="thin", color="FF000000"),
#                 top=Side(border_style="thin", color="FF000000"),
#                 bottom=Side(border_style="thin", color="FF000000"),
#             )
#             # set width
#             worksheet.column_dimensions[alphabetDict[j]].width = 20
#             if i == 1:
#                 # set col name
#                 worksheet[f"{alphabetDict[j]}{i}"].value = col_names[j - 1]
#             else:
#                 # set value
#                 worksheet[f"{alphabetDict[j]}{i}"].value = value_dict[j]
#                 if j == 6:
#                     # set cell style
#                     worksheet[f"{alphabetDict[j]}{i}"].number_format = "#,##0.00"
#                     # set to right
#                     worksheet[f"{alphabetDict[j]}{i}"].alignment = Alignment(
#                         horizontal="right", vertical="center"
#                     )
#                 if j == 4 and value_dict[4]:
#                     # set cell style
#                     worksheet[f"{alphabetDict[j]}{i}"].number_format = "#,##0.00"
#                     # set font color as red
#                     worksheet[f"{alphabetDict[j]}{i}"].font = Font(
#                         name="微軟正黑體", size=12, color="FF0000"
#                     )
#                     # set to right
#                     worksheet[f"{alphabetDict[j]}{i}"].alignment = Alignment(
#                         horizontal="right", vertical="center"
#                     )
#                 if j == 5 and value_dict[5]:
#                     # set cell style
#                     worksheet[f"{alphabetDict[j]}{i}"].number_format = "#,##0.00"
#                     # set font bg color as blue
#                     worksheet[f"{alphabetDict[j]}{i}"].fill = PatternFill(
#                         fill_type="solid", start_color="00b0f0", end_color="00b0f0"
#                     )
#                     # set to right
#                     worksheet[f"{alphabetDict[j]}{i}"].alignment = Alignment(
#                         horizontal="right", vertical="center"
#                     )
#     # 合併A-E
#     worksheet.merge_cells(f"A{max_rows + 2}:E{max_rows + 2}")
#     worksheet[f"A{max_rows + 2}"].value = "Sub-total Balance"
#     # 靠右側
#     worksheet[f"A{max_rows + 2}"].alignment = Alignment(
#         horizontal="right", vertical="center"
#     )
#     # set font color as blue and 粗體
#     worksheet[f"A{max_rows + 2}"].font = Font(
#         name="微軟正黑體", size=12, color="0000FF", bold=True
#     )
#
#     worksheet[f"F{max_rows + 2}"].value = OriginalCreditBalance
#     # set cell style
#     worksheet[f"F{max_rows + 2}"].number_format = "#,##0.00"
#     # set font color as blue and 粗體
#     worksheet[f"F{max_rows + 2}"].font = Font(
#         name="微軟正黑體", size=12, color="0000FF", bold=True
#     )
#     worksheet[f"F{max_rows + 2}"].alignment = Alignment(
#         horizontal="right", vertical="center"
#     )
#
#     # set file
#     file_name = f"CB歷程.xlsx"
#     workbook.save(file_name)
#
#     # get FastAPI project path
#     project_path = os.path.dirname(
#         os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
#     )
#     print(project_path)
#
#     # generate StreamingResponse
#     file_path = os.path.join(project_path, file_name)
#     headers = {"Content-Disposition": 'attachment; filename="CB Record.xlsx"'}
#     return FileResponse(file_path, headers=headers)


@router.post("/CreditBalanceStatement/generateReport")
async def generateReport(
    request: Request,
    db: Session = Depends(get_db),
):
    """
    {
        "CBID": 1
    }
    """
    crudCreditBalance = CRUD(db, CreditBalanceDBModel)
    crudCreditBalanceStatement = CRUD(db, CreditBalanceStatementDBModel)
    crudBillDetail = CRUD(db, BillDetailDBModel)
    crudBillMaster = CRUD(db, BillMasterDBModel)
    crudCreditNote = CRUD(db, CreditNoteDBModel)
    crudCreditNoteDetail = CRUD(db, CreditNoteDetailDBModel)

    CBData = crudCreditBalance.get_with_condition(
        {"CBID": (await request.json())["CBID"]}
    )[0]
    CBStatementDataList = crudCreditBalanceStatement.get_with_condition(
        {"CBID": (await request.json())["CBID"]}
    )

    SubmarineCableList = []
    WorkTitleList = []
    BillMilestoneList = []
    InvNoList = []
    BillIssueDateList = []
    CNNoList = []
    CNIssueDateList = []
    DescriptionList = []
    DebitList = []
    CreditList = []
    BalanceList = []

    for i, CBStatementData in enumerate(CBStatementDataList):
        pprint(orm_to_dict(CBStatementData))
        if CBStatementData.TransItem == "BM_ADD":
            BillMilestoneList.append(CBData.BillMilestone)
            InvNoList.append(CBStatementData.BillingNo)

            # get BillMaster
            BillDetailData = crudBillDetail.get_with_condition(
                {"BillDetailID": CBStatementData.BLDetailID}
            )[0]
            BillMasterData = crudBillMaster.get_with_condition(
                {"BillMasterID": BillDetailData.BillMasterID}
            )[0]

            SubmarineCableList.append(CBData.SubmarineCable)
            WorkTitleList.append(CBData.WorkTitle)

            # append BillMaster's IssueDate
            BillIssueDateList.append(orm_to_dict(BillMasterData)["IssueDate"])

            # BM_ADD no CNNo
            CNNoList.append("")
            CNIssueDateList.append("")

            # append BillDetail's FeeItem
            DescriptionList.append(BillDetailData.FeeItem)

            # No Debit
            DebitList.append("")

            CreditList.append(CBStatementData.OrgAmount)
            BalanceList.append(CBData.CurrAmount)
        elif CBStatementData.TransItem == "USER_ADD":
            BillMilestoneList.append(
                CBData.BillMilestone if CBData.BillMilestone else ""
            )
            SubmarineCableList.append(
                CBData.SubmarineCable if CBData.SubmarineCable else ""
            )
            WorkTitleList.append(CBData.WorkTitle if CBData.WorkTitle else "")

            InvNoList.append(
                CBStatementData.BillingNo if CBStatementData.BillingNo else ""
            )
            BillIssueDateList.append(orm_to_dict(CBStatementData)["CreateDate"])

            # get CNNo
            CNDetailData = crudCreditNoteDetail.get_with_condition(
                {"CBStateID": CBStatementData.CBStateID}
            )
            CNNoList.append(CBData.CNNo if CBData.CNNo else "")

            # get CN IssueDate
            if CNDetailData:
                CNData = crudCreditNote.get_with_condition({"CNID": CNDetailData.CNID})
                CNIssueDateList.append(orm_to_dict(CNData[0]).IssueDate)
            else:
                CNIssueDateList.append("")

            DescriptionList.append(CBStatementData.Note if CBStatementData.Note else "")
            DebitList.append("")
            CreditList.append(CBStatementData.OrgAmount)
            BalanceList.append(CBStatementData.OrgAmount)
        elif CBStatementData.TransItem == "RETURN":
            BillMilestoneList.append(
                CBData.BillMilestone if CBData.BillMilestone else ""
            )
            InvNoList.append(CBData.BillingNo if CBData.BillingNo else "")
            SubmarineCableList.append(CBData.SubmarineCable)
            WorkTitleList.append(CBData.WorkTitle)

            # get BillMaster
            BillDetailData = crudBillDetail.get_with_condition(
                {"BillDetailID": CBStatementData.BLDetailID}
            )[0]
            BillMasterData = crudBillDetail.get_with_condition(
                {"BillMasterID": BillDetailData.BillMasterID}
            )[0]
            BillIssueDateList.append(orm_to_dict(BillMasterData)["IssueDate"])

            # get CN info
            CNDetailData = crudCreditNoteDetail.get_with_condition(
                {"CBStateID": CBStatementData.CBStateID}
            )
            CNNoList.append(CBData.CNNo if CBData.CNNo else "")

            # get CN IssueDate
            if CNDetailData:
                CNData = crudCreditNote.get_with_condition({"CNID": CNDetailData.CNID})
                CNIssueDateList.append(orm_to_dict(CNData[0]).IssueDate)
            else:
                CNIssueDateList.append("")

            DescriptionList.append(
                BillDetailData.FeeItem if BillDetailData.FeeItem else ""
            )
            DebitList.append(
                abs(CBStatementData.TransAmount) if CBStatementData.TransAmount else ""
            )
            CreditList.append("")
            BalanceList.append(CBStatementData.OrgAmount + CBStatementData.TransAmount)
        elif CBStatementData.TransItem == "DEDUCT":
            BillDetailData = crudBillDetail.get_with_condition(
                {"BillDetailID": CBStatementData.BLDetailID}
            )[0]
            BillMasterData = crudBillMaster.get_with_condition(
                {"BillMasterID": BillDetailData.BillMasterID}
            )[0]
            SubmarineCableList.append(CBData.SubmarineCable)
            BillMilestoneList.append(
                BillDetailData.BillMilestone if BillDetailData.BillMilestone else ""
            )
            WorkTitleList.append(
                BillDetailData.WorkTitle if BillDetailData.WorkTitle else ""
            )
            InvNoList.append(
                BillMasterData.BillingNo if BillMasterData.BillingNo else ""
            )
            BillIssueDateList.append(orm_to_dict(BillMasterData)["IssueDate"])
            CNNoList.append("")
            CNIssueDateList.append("")
            DescriptionList.append(
                BillDetailData.FeeItem if BillDetailData.FeeItem else ""
            )
            DebitList.append(
                abs(CBStatementData.TransAmount) if CBStatementData.TransAmount else ""
            )
            CreditList.append("")
            BalanceList.append(CBStatementData.OrgAmount + CBStatementData.TransAmount)
        elif CBStatementData.TransItem == "REFUND":
            BillDetailData = crudBillDetail.get_with_condition(
                {"BillDetailID": CBStatementData.BLDetailID}
            )[0]
            BillMasterData = crudBillMaster.get_with_condition(
                {"BillMasterID": BillDetailData.BillMasterID}
            )[0]
            SubmarineCableList.append(
                BillDetailData.SubmarineCable if BillDetailData.SubmarineCable else ""
            )
            WorkTitleList.append(
                BillDetailData.WorkTitle if BillDetailData.WorkTitle else ""
            )
            BillMilestoneList.append(
                BillDetailData.BillMilestone if BillDetailData.BillMilestone else ""
            )
            InvNoList.append(
                BillMasterData.BillingNo if BillMasterData.BillingNo else ""
            )
            BillIssueDateList.append(orm_to_dict(BillMasterData)["IssueDate"])

            # get CN info
            CNDetailData = crudCreditNoteDetail.get_with_condition(
                {"CBStateID": CBStatementData.CBStateID}
            )
            CNNoList.append(CBData.CNNo if CBData.CNNo else "")

            # get CN IssueDate
            if CNDetailData:
                CNData = crudCreditNote.get_with_condition({"CNID": CNDetailData.CNID})
                CNIssueDateList.append(orm_to_dict(CNData[0]).IssueDate)
            else:
                CNIssueDateList.append("")

            DescriptionList.append(
                BillDetailData.FeeItem if BillDetailData.FeeItem else ""
            )
            DebitList.append(
                abs(CBStatementData.TransAmount) if CBStatementData.TransAmount else ""
            )
            CreditList.append("")
            BalanceList.append(CBStatementData.OrgAmount + CBStatementData.TransAmount)
        elif CBStatementData.TransItem == "BANK_FEE":
            BillDetailData = crudBillDetail.get_with_condition(
                {"BillDetailID": CBStatementData.BLDetailID}
            )[0]
            BillMasterData = crudBillMaster.get_with_condition(
                {"BillMasterID": BillDetailData.BillMasterID}
            )[0]
            SubmarineCableList.append(
                BillDetailData.SubmarineCable if BillDetailData.SubmarineCable else ""
            )
            WorkTitleList.append(
                BillDetailData.WorkTitle if BillDetailData.WorkTitle else ""
            )
            BillMilestoneList.append(
                BillDetailData.BillMilestone if BillDetailData.BillMilestone else ""
            )
            BillIssueDateList.append(orm_to_dict(BillMasterData)["IssueDate"])

            # get CNNo
            CNDetailData = crudCreditNoteDetail.get_with_condition(
                {"CBStateID": CBStatementData.CBStateID}
            )
            CNNoList.append(CBData.CNNo if CBData.CNNo else "")

            # get CN IssueDate
            if CNDetailData:
                CNData = crudCreditNote.get_with_condition({"CNID": CNDetailData.CNID})
                CNIssueDateList.append(orm_to_dict(CNData[0]).IssueDate)
            else:
                CNIssueDateList.append("")

            DescriptionList.append(
                BillDetailData.FeeItem if BillDetailData.FeeItem else ""
            )
            DebitList.append(
                abs(CBStatementData.TransAmount) if CBStatementData.TransAmount else ""
            )
            CreditList.append("")
            BalanceList.append(CBStatementData.OrgAmount + CBStatementData.TransAmount)

    dict_data = {
        "SubmarineCable": SubmarineCableList,
        "WorkTitle": WorkTitleList,
        "BillMilestone": BillMilestoneList,
        "InvNo": InvNoList,
        "BillIssueDate": BillIssueDateList,
        "CNNo": CNNoList,
        "CNIssueDate": CNIssueDateList,
        "Description": DescriptionList,
        "Debit": DebitList,
        "Credit": CreditList,
        "Balance": BalanceList,
    }
    df = pd.DataFrame(dict_data)

    # save to excel
    # 時間戳
    timestamp = convert_time_to_str(datetime.now(), "%Y%m%d%H%M%S")
    df.to_excel(f"CreditBalanceReport_{timestamp}.xlsx", index=False)

    return dict_data


@router.post("/CreditBalance", status_code=status.HTTP_201_CREATED)
async def addCreditBalance(
    request: Request,
    db: Session = Depends(get_db),
):
    CreditBalanceDictData = await request.json()
    if isinstance(CreditBalanceDictData["CurrAmount"], str):
        CreditBalanceDictData["CurrAmount"] = float(
            CreditBalanceDictData["CurrAmount"].replace(",", "")
        )

    crudCreditBalance = CRUD(db, CreditBalanceDBModel)
    crudCreditBalanceStatement = CRUD(db, CreditBalanceStatementDBModel)
    crudCreditNote = CRUD(db, CreditNoteDBModel)
    crudCreditNoteDetail = CRUD(db, CreditNoteDetailDBModel)
    crudParties = CRUD(db, PartiesDBModel)
    crudSubmarineCables = CRUD(db, SubmarineCablesDBModel)

    PartyCode = crudParties.get_with_condition(
        {"PartyName": CreditBalanceDictData["PartyName"]}
    )[0].PartyCode

    CableCode = crudSubmarineCables.get_with_condition(
        {"CableName": CreditBalanceDictData["SubmarineCable"]}
    )[0].CableCode

    timestamp = (
        convert_time_to_str(datetime.now())
        .replace(" ", "")
        .replace(":", "")
        .replace("-", "")[2:12]
    )
    WorkTitleMapping = {"Upgrade": "UP", "Construction": "CO", "O&M": "OM"}

    # 新增CB
    CreditBalanceDictData["CreateDate"] = convert_time_to_str(datetime.now())
    CreditBalanceDictData[
        "CNNo"
    ] = f"CN{CableCode}{WorkTitleMapping[CreditBalanceDictData['WorkTitle']]}-{PartyCode}{timestamp}"
    CreditBalancePydanticData = CreditBalanceSchema(**CreditBalanceDictData)
    newCreditBalanceData = crudCreditBalance.create(CreditBalancePydanticData)

    # 新增CBStatement
    CBStatementDictData = {
        "CBID": newCreditBalanceData.CBID,
        "BillingNo": newCreditBalanceData.BillingNo,
        "TransItem": "USER_ADD",
        "OrgAmount": newCreditBalanceData.CurrAmount,
        "TransAmount": 0,
        "Note": newCreditBalanceData.Note,
        "CreateDate": convert_time_to_str(datetime.now()),
    }

    newCBStatementData = crudCreditBalanceStatement.create(
        CreditBalanceStatementSchema(**CBStatementDictData)
    )

    # # 新增CN
    # newCNDictData = {
    #     "CNNo": CreditBalanceDictData["CNNo"],
    #     "CNType": CBStatementDictData["TransItem"],
    #     "SubmarineCable": CreditBalanceDictData["SubmarineCable"],
    #     "WorkTitle": CreditBalanceDictData["WorkTitle"],
    #     "PartyName": CreditBalanceDictData["PartyName"],
    #     "CurrAmount": CreditBalanceDictData["CurrAmount"],
    #     "IssyeDate": convert_time_to_str(datetime.now()),
    #     "Note": "",
    #     "URI": ""
    # }

    return {
        "message": "CreditBalance successfully created",
        "CreditBalance": newCreditBalanceData,
        "CBStatement": newCBStatementData,
    }


@router.post("/updateCreditBalance", status_code=status.HTTP_200_OK)
async def updateCreditBalance(
    request: Request,
    db: Session = Depends(get_db),
):
    CBDictData = await request.json()
    crud = CRUD(db, CreditBalanceDBModel)
    CBData = crud.get_with_condition({"CBID": CBDictData["CBID"]})[0]
    newCBData = crud.update(CBData, CBDictData)
    return {"message": "CreditBalance successfully updated", "newCBData": newCBData}
